"""
Excelファイルをパースして排出量データを抽出
"""
import json
import csv
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# 設定
DATA_DIR = Path(__file__).parent.parent / "data"
RAW_DIR = DATA_DIR / "raw"
PROCESSED_DIR = DATA_DIR / "processed"
MUNICIPALITIES_CSV = DATA_DIR / "tokyo_municipalities.csv"

# Excelシート設定
SHEET_NAME = "データシート1"
HEADER_ROW = 8  # 8行目がヘッダー
DATA_START_ROW = 9  # 9行目からデータ

# 列インデックス（1ベース）
COL_YEAR = 5  # 西暦
COL_CITY_CODE = 7  # 団体コード
COL_CITY_NAME = 8  # 自治体名
COL_DATA_START = 10  # データ開始列（製造業から）


def parse_excel(file_path: Path, city_code: str, city_name: str) -> Optional[Dict]:
    """
    Excelファイルをパースして排出量データを抽出

    Args:
        file_path: Excelファイルパス
        city_code: 団体コード
        city_name: 自治体名

    Returns:
        パース結果の辞書、エラー時はNone
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        if SHEET_NAME not in wb.sheetnames:
            print(f"[ERROR] {city_code} - シート '{SHEET_NAME}' が見つかりません")
            return None

        ws: Worksheet = wb[SHEET_NAME]

        # ヘッダー行から部門名を取得
        # Column 10-19のみ（排出量データ）を読み取る
        sectors = {}
        for col in range(COL_DATA_START, COL_DATA_START + 10):  # 10-19列目のみ
            header = ws.cell(HEADER_ROW, col).value
            if header:
                # "aa_製造業" -> "製造業" のように変換
                sector_name = str(header).replace('aa_', '').replace('部門', '')
                # 簡易的な部門名マッピング
                if '製造' in sector_name:
                    sector_name = '製造業'
                elif '建設' in sector_name or '鉱業' in sector_name:
                    sector_name = '建設業'
                elif '農林' in sector_name or '農業' in sector_name:
                    sector_name = '農林水産業'
                elif '業務' in sector_name:
                    sector_name = '業務その他'
                elif '家庭' in sector_name:
                    sector_name = '家庭'
                elif '旅客' in sector_name or '自動車' in sector_name:
                    sector_name = '旅客'
                elif '貨物' in sector_name:
                    sector_name = '貨物'
                elif '鉄道' in sector_name:
                    sector_name = '鉄道'
                elif '船舶' in sector_name:
                    sector_name = '船舶'
                elif '廃棄' in sector_name:
                    sector_name = '廃棄物'
                sectors[col] = sector_name

        # データ行を走査
        emissions_by_sector = {}
        years_set = set()

        for row in range(DATA_START_ROW, ws.max_row + 1):
            # 団体コードが一致する行のみ処理
            row_city_code = ws.cell(row, COL_CITY_CODE).value
            if row_city_code and str(row_city_code) == str(city_code):
                year_val = ws.cell(row, COL_YEAR).value
                if year_val:
                    try:
                        year = int(year_val)
                        years_set.add(year)

                        # 各部門のデータを取得
                        for col, sector_name in sectors.items():
                            value = ws.cell(row, col).value
                            if value is not None:
                                try:
                                    emission_value = float(value)
                                    if sector_name not in emissions_by_sector:
                                        emissions_by_sector[sector_name] = {}
                                    emissions_by_sector[sector_name][year] = emission_value
                                except (ValueError, TypeError):
                                    pass
                    except (ValueError, TypeError):
                        pass

        if not years_set:
            print(f"[ERROR] {city_code} - データが見つかりません")
            return None

        years = sorted(list(years_set))

        # 結果を構築
        result = {
            "city_code": city_code,
            "city_name": city_name,
            "years": years,
            "emissions": emissions_by_sector
        }

        return result

    except Exception as e:
        print(f"[ERROR] {city_code} - パース失敗: {e}")
        return None


def main():
    """メイン処理"""
    # 出力ディレクトリを作成
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    # CSVから自治体情報を読み込み
    municipalities = []
    with open(MUNICIPALITIES_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            municipalities.append({
                'city_code': row['city_code'],
                'name': row['name'],
                'region': row['region']
            })

    print(f"東京都 {len(municipalities)} 自治体のExcelファイルをパース開始")
    print("-" * 60)

    success_count = 0
    fail_count = 0
    all_data = []

    for i, muni in enumerate(municipalities, 1):
        city_code = muni['city_code']
        city_name = muni['name']
        excel_path = RAW_DIR / f"{city_code}.xlsx"

        print(f"[{i}/{len(municipalities)}] {city_name} ({city_code})... ", end="", flush=True)

        if not excel_path.exists():
            print(f"[SKIP] Excelファイルが存在しません")
            fail_count += 1
            continue

        result = parse_excel(excel_path, city_code, city_name)

        if result:
            all_data.append(result)
            success_count += 1
            print(f"[OK] {len(result['years'])}年度分")
        else:
            fail_count += 1

    # JSONファイルとして保存
    output_path = PROCESSED_DIR / "tokyo_emissions.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    print("-" * 60)
    print(f"完了: 成功 {success_count} / 失敗 {fail_count} / 合計 {len(municipalities)}")
    print(f"出力: {output_path}")


if __name__ == "__main__":
    main()
